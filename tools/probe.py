#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Survey a folder of raw records and say how each file should be read.

    python3 tools/probe.py workspace/inbox
    python3 tools/probe.py workspace/inbox --json      # machine-readable

Run this before writing a single line of parsing. The one question that decides your
whole approach is *does this PDF carry a text layer* — if it does, PyMuPDF gives you
the exact characters for free; if it doesn't, you (the agent) look at the rendered page
yourself. This tool answers that per file and prints the follow-up command for each.

It reports, never modifies. See docs/20-extract.md for what to do with the answers.
"""
import argparse, json, os, sys, zipfile

TEXT_LAYER_MIN = 120      # chars/page below this reads as "scanned", not "digital"
DECOR_MAX_PX = 160        # embedded images smaller than this are logos / QR / signatures


def probe_pdf(path):
    import fitz
    info = {'kind': 'pdf', 'pages': 0, 'chars': 0, 'per_page': [], 'images': [],
            'encrypted': False, 'error': None}
    try:
        doc = fitz.open(path)
    except Exception as e:
        info['error'] = f'{type(e).__name__}: {e}'
        return info
    if doc.needs_pass:
        info['encrypted'] = True
        doc.close()
        return info
    info['pages'] = doc.page_count
    for page in doc:
        n = len(page.get_text().strip())
        info['per_page'].append(n)
        info['chars'] += n
        for img in page.get_images(full=True):
            try:
                pix = doc.extract_image(img[0])
                info['images'].append((pix['width'], pix['height']))
            except Exception:
                pass
    doc.close()
    per = info['chars'] / max(1, info['pages'])
    info['layer'] = 1 if per >= TEXT_LAYER_MIN else 2
    info['diagnostic_images'] = [wh for wh in info['images']
                                 if max(wh) > DECOR_MAX_PX]
    return info


def probe_image(path):
    from PIL import Image
    try:
        with Image.open(path) as im:
            return {'kind': 'image', 'size': im.size, 'layer': 2, 'error': None}
    except Exception as e:
        return {'kind': 'image', 'layer': 2, 'error': f'{type(e).__name__}: {e}'}


def probe_xlsx(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = [(ws.title, ws.max_row, ws.max_column) for ws in wb.worksheets]
        wb.close()
        return {'kind': 'xlsx', 'sheets': sheets, 'layer': 1, 'error': None}
    except Exception as e:
        return {'kind': 'xlsx', 'layer': 1, 'error': f'{type(e).__name__}: {e}'}


def probe_zip(path):
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()
        garbled = sum(1 for n in names if any(0xE000 <= ord(c) <= 0xF8FF or
                                              'À' <= c <= 'ÿ' for c in n))
        return {'kind': 'zip', 'entries': len(names), 'garbled_names': garbled,
                'layer': 0, 'error': None}
    except Exception as e:
        return {'kind': 'zip', 'layer': 0, 'error': f'{type(e).__name__}: {e}'}


IMG_EXT = {'.png', '.jpg', '.jpeg', '.webp', '.heic', '.tif', '.tiff', '.bmp'}
VID_EXT = {'.mp4', '.mov', '.m4v', '.avi', '.mkv'}


def probe_one(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.pdf':
        return probe_pdf(path)
    if ext in IMG_EXT:
        return probe_image(path)
    if ext in ('.xlsx', '.xlsm'):
        return probe_xlsx(path)
    if ext == '.zip':
        return probe_zip(path)
    if ext in VID_EXT:
        return {'kind': 'video', 'layer': 2, 'error': None}
    return {'kind': ext.lstrip('.') or 'file', 'layer': None, 'error': None}


def walk(root):
    out = []
    for dirpath, _, files in os.walk(root):
        for f in sorted(files):
            if f.startswith('.'):
                continue
            p = os.path.join(dirpath, f)
            rec = probe_one(p)
            rec['path'] = os.path.relpath(p, root)
            rec['bytes'] = os.path.getsize(p)
            out.append(rec)
    return out


def report(root, recs):
    if not recs:
        print(f'{root} 是空的 —— 先把报告文件放进去（见 docs/12-acquire-manual.md）')
        return 1

    l1 = [r for r in recs if r.get('layer') == 1]
    l2 = [r for r in recs if r.get('layer') == 2]
    zips = [r for r in recs if r['kind'] == 'zip']
    vids = [r for r in recs if r['kind'] == 'video']
    enc = [r for r in recs if r.get('encrypted')]
    err = [r for r in recs if r.get('error')]

    print(f'{root}：{len(recs)} 个文件，{sum(r["bytes"] for r in recs) / 1048576:.1f} MB\n')
    for r in recs:
        tag = {'pdf': 'PDF', 'image': '图片', 'xlsx': 'Excel', 'zip': '压缩包',
               'video': '视频'}.get(r['kind'], r['kind'])
        line = f"  {r['path']}  [{tag}]"
        if r['kind'] == 'pdf' and not r.get('error'):
            per = r['chars'] / max(1, r['pages'])
            line += (f"  {r['pages']}页 · 文字层 {r['chars']} 字（{per:.0f}/页）"
                     f" · 内嵌图 {len(r['images'])} 张"
                     f"（够大可能是诊断图的 {len(r['diagnostic_images'])} 张）")
            line += '  → 第1层：直接取文字' if r['layer'] == 1 else '  → 第2层：渲染成图，你自己读'
        elif r['kind'] == 'image':
            line += f"  {r.get('size')}  → 第2层：你自己读"
        elif r['kind'] == 'xlsx' and not r.get('error'):
            line += '  ' + '；'.join(f'{n} {rows}×{cols}' for n, rows, cols in r['sheets'])
        elif r['kind'] == 'zip' and not r.get('error'):
            line += f"  {r['entries']} 项" + (f"（{r['garbled_names']} 个文件名疑似乱码）"
                                              if r['garbled_names'] else '')
        if r.get('encrypted'):
            line += '  ⚠ 有密码，需要用户提供'
        if r.get('error'):
            line += f"  ⚠ 打不开：{r['error']}"
        print(line)

    print(f'\n分流：第1层 {len(l1)} 个（有文字层/表格，直接取）·'
          f' 第2层 {len(l2)} 个（要你看图读）')
    print('\n下一步：')
    if zips:
        print("  # 先解压（自动修 Windows 中文文件名乱码）")
        print("  python3 -c \"import sys;sys.path.insert(0,'tools');from lib import unzip_cjk;"
              f"unzip_cjk('{os.path.join(root, zips[0]['path'])}','workspace/raw')\"")
    if l1:
        print(f"  python3 tools/extract.py text {os.path.join(root, l1[0]['path'])}"
              "        # 逐份取文字，零误差")
    if l2:
        print(f"  python3 tools/extract.py pages {os.path.join(root, l2[0]['path'])}"
              "       # 渲染成 PNG，然后用 Read 打开图片自己读")
    if vids:
        print("  ffmpeg -i 录屏.mp4 -vf \"fps=2,mpdecimate\" -vsync vfr "
              "workspace/raw/frame_%04d.png    # 抽帧去重后同第2层")
    if len(l2) > 12:
        print(f"  # 第2层有 {len(l2)} 个，可能撑爆上下文 → 可选：python3 tools/fanout.py --help")
    if enc:
        print(f"  ⚠ {len(enc)} 个 PDF 有密码。请用户自己解密后再放进来——不要去猜密码。")
    return 1 if err else 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('root', nargs='?', default='workspace/inbox')
    ap.add_argument('--json', action='store_true')
    a = ap.parse_args()
    if not os.path.isdir(a.root):
        sys.exit(f'没有这个目录：{a.root}')
    recs = walk(a.root)
    if a.json:
        print(json.dumps(recs, ensure_ascii=False, indent=1))
        return 0
    return report(a.root, recs)


if __name__ == '__main__':
    sys.exit(main())
